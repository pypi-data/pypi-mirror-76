from openpyxl import workbook
from openpyxl import load_workbook
from robot.api import logger
from SeleniumLibrary.base import keyword, LibraryComponent
from robot.libraries.BuiltIn import BuiltIn, run_keyword_variant
import json


class BDSExcelDriver:
    def __init__(self, file=None, sheet_name='Sheet1'):
        self.file = file
        self.sheet_name = sheet_name
        self.builtIn = BuiltIn()

    @keyword
    def _write_to_existing_excel(self, value, column='A', row='1'):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(self.sheet_name) == 0):
            self.builtIn.fail('Đường dẫn không chính xác')
        sheet = wb[self.sheet_name]
        sheet[column + row] = value
        wb.save(self.file)

    @keyword
    def _read_from_existing_excel(self, column='A', row='1'):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(self.sheet_name) == 0):
            self.builtIn.fail('Đường dẫn không chính xác')
        sheet = wb[self.sheet_name]
        result = sheet[column + row].value
        if(result == '${EMPTY}'):
            result = ''
        return result

    @keyword
    def set_global_value_from_existing_excel(self, *value_names):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(self.sheet_name) == 0):
            self.builtIn.fail('Đường dẫn không chính xác')
        sheet = wb[self.sheet_name]
        test_case_column = None
        value_row = None
        row = None
        value_columns = {}
        test_name = self.builtIn._get_var_value('${TEST_NAME}', '')
        for rows in tuple(sheet.rows):
            for cell in rows:
                if cell.value == '*** Test Cases ***':
                    test_case_column = cell.column_letter
                    value_row = cell.row

                if test_case_column != None and cell.column_letter == test_case_column and cell.value == test_name:
                    row = cell.row

                if value_row != None and cell.row == value_row and cell.column_letter != '*** Test Cases ***':
                    value_columns[cell.value] = cell.column_letter
                    
        for value_name in value_names:
            name = '${' + value_name + '}'
            if name in value_columns:
                self.builtIn.set_global_variable(name, sheet[value_columns[name] + str(row)].value)
    
    @keyword
    def set_suite_value_from_existing_excel(self, *value_names):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(self.sheet_name) == 0):
            self.builtIn.fail('Đường dẫn không chính xác')
        sheet = wb[self.sheet_name]
        test_case_column = None
        value_row = None
        row = None
        value_columns = {}
        test_name = self.builtIn._get_var_value('${TEST_NAME}', '')
        for rows in tuple(sheet.rows):
            for cell in rows:
                if cell.value == '*** Test Cases ***':
                    test_case_column = cell.column_letter
                    value_row = cell.row

                if test_case_column != None and cell.column_letter == test_case_column and cell.value == test_name:
                    row = cell.row

                if value_row != None and cell.row == value_row and cell.column_letter != '*** Test Cases ***':
                    value_columns[cell.value] = cell.column_letter
                    
        for value_name in value_names:
            name = '${' + value_name + '}'
            if name in value_columns:
                self.builtIn.set_suite_variable(name, sheet[value_columns[name] + str(row)].value)
        
    @keyword        
    def set_test_value_from_existing_excel(self, *value_names):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(self.sheet_name) == 0):
            self.builtIn.fail('Đường dẫn không chính xác')
        sheet = wb[self.sheet_name]
        test_case_column = None
        value_row = None
        row = None
        value_columns = {}
        test_name = self.builtIn._get_var_value('${TEST_NAME}', '')
        for rows in tuple(sheet.rows):
            for cell in rows:
                if cell.value == '*** Test Cases ***':
                    test_case_column = cell.column_letter
                    value_row = cell.row

                if test_case_column != None and cell.column_letter == test_case_column and cell.value == test_name:
                    row = cell.row

                if value_row != None and cell.row == value_row and cell.column_letter != '*** Test Cases ***':
                    value_columns[cell.value] = cell.column_letter
                    
        for value_name in value_names:
            name = '${' + value_name + '}'
            if name in value_columns:
                self.builtIn.set_test_variable(name, sheet[value_columns[name] + str(row)].value)

        
